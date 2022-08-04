# This module defines the class that represents a single budget. Takes in a
# parsed Config object and does the rest of the work.
#
#   Connor Shugg

# Imports
import os
import sys
from datetime import datetime
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import LineChart, Reference, Series

# Enable import from the parent directory
dpath = os.path.dirname(os.path.realpath(__file__)) # directory of this file
dpath = os.path.dirname(dpath)                      # parent directory
if dpath not in sys.path:                           # add to path
        sys.path.append(dpath)

# Local imports
from lib.bclass import BudgetClass, BudgetClassType
from lib.transaction import Transaction

# Simple class used to represent a return value from these 
class BudgetResult:
    # Constructor. Sets up fields.
    def __init__(self, success=True, msg="", data=None):
        self.success = success
        self.message = msg
        self.data = data

# Budget class
class Budget:
    # Takes in the Config object that was parsed prior to this object's
    # creation.
    def __init__(self, conf):
        self.conf = conf
        self.classes = []
        self.savings = conf.surplus_savings
        self.reset_dates = conf.reset_dates

        # before we load the classes, we need to know if today is a reset date
        # for the budget. If it is, we'll delete all transactions and start
        # saving to a new backup location
        nrd = self.conf.reset_dates[0]
        now = datetime.now()
        today_is_reset = nrd.month == now.month and nrd.day == now.day

        # get the current save root path based on the current reset date
        sroot = self.save_root_path()

        try:
            # if today is a reset day, we'll back up the config file itself to the
            # new backup location
            bdpath = self.backup_setup()
            if today_is_reset:
                shutil.copy(self.conf.fpath, os.path.join(bdpath, "config.json"))
        except Exception as e:
            # if we fail to setup the backup location, don't panic
            pass

        # iterate through the previous period's directory's files to load in
        # each budget class
        day = datetime.fromtimestamp(now.timestamp() - 86400)
        for root, dirs, files in os.walk(self.save_root_path(dt=day)):
            for f in files:
                # if the file has the JSON extension, we'll try to load it as a
                # budget class object
                if f.lower().endswith(".json") and "config" not in f.lower():
                    # if the class exists in *today's* save path, load that
                    # instead. If it doesn't, we'll use yesterday's. This
                    # prevents multiple resets on the same reset day.
                    lpath = os.path.join(root, f)
                    if os.path.isfile(os.path.join(sroot, f)):
                        lpath = os.path.join(sroot, f)
                    bc = BudgetClass.load(lpath)
                    self.classes.append(bc)
                    
                    # if today is a reset day AND the class's last reset date
                    # was before today (meaning we haven't yet reset this class
                    # for this cycle), remove all non-recurring transactions and
                    # write the budget class back out to disk
                    class_needs_reset = bc.last_reset == None or \
                                        bc.last_reset.month != nrd.month or \
                                        bc.last_reset.day != nrd.day or \
                                        bc.last_reset.year != nrd.year
                    if today_is_reset and class_needs_reset:
                        bc.reset()
                        # save the class to the *new* location
                        bc.save(os.path.join(sroot, f))
        try:
            # attempt to initialize the backup location, and save all budget classes
            # to the backup location if they don't exist
            for bc in self.classes:
                fpath = self.backup_class_path(bc)
                if not os.path.isfile(fpath):
                    bc.save(self.backup_class_path(bc))
        except Exception as e:
            # if we fail to set up the backup location, don't panic
            pass
    
    # Used to iterate through the budget's classes.
    def __iter__(self):
        for bc in self.all():
            yield bc

    # ------------------------------ Additions ------------------------------- #
    # Takes in a BudgetClass object and adds it to the budget. Upon calling this
    # function, the budget is written out to a file.
    def add_class(self, bclass):
        # before appending to the array, make sure there are no name conflicts
        for bc in self.classes:
            assert bclass.name.lower() != bc.name.lower(), \
                   "duplicate budget class name detected"
        self.classes.append(bclass)

        # write out to a file
        sroot = self.save_root_path()
        fpath = os.path.join(sroot, bclass.to_file_name())
        bclass.save(fpath)
        # attempt to back up
        try:
            bclass.save(self.backup_class_path(bclass))
        except Exception as e:
            m = "Failed to backup class: %s" % e
            return BudgetResult(success=True, msg=m)
        return BudgetResult(success=True)
    
    # Takes in a transaction and a budget class and adds it to the budget class,
    # then saves the budget class out to disk.
    def add_transaction(self, bclass, transaction):
        # get the save path based on the transaction's timestamp. It's possible
        # this will *not* be in the current budget period (i.e. after some other
        # reset date). We'll check for this here and react accordingly
        sroot = self.save_root_path(dt=transaction.timestamp)
        fpath = os.path.join(sroot, bclass.to_file_name())
        if (sroot != self.save_root_path()): # if transaction path != the current path
            # in this case, we need to use the version of the budget class saved
            # in the computed path (i.e., for a different reset date). To do this,
            # we'll try to load in the file, if it exists
            if os.path.isfile(fpath):
                bc = None
                try:
                    bc = BudgetClass.load(fpath)
                    bc.add(transaction)
                    bc.save(fpath)
                except Exception as e:
                    m = "Failed to update budget class from different reset date: %s" % e
                    return BudgetResult(success=False, msg=m)
            else:
                # if one doesn't exist yet for this different reset date, we'll
                # create a copy of the current budget class, with this transaction
                # as the only transaction in its history
                bc = bclass.copy()
                bc.history = []
                bc.add(transaction)
                bc.save(fpath)
        else:
            # if we're not saving to a different reset date within the budget,
            # things are much easier - simply add and save the current class
            bclass.add(transaction)
            bclass.save(fpath)

        # attempt to back up the class we just saved
        try:
            bclass.save(self.backup_class_path(bclass, dt=transaction.timestamp))
        except Exception as e:
            m = "Failed to backup class: %s" % e
            return BudgetResult(success=True, msg=m)
        return BudgetResult(success=True)


    # ------------------------------ Searching ------------------------------- #
    # Expects a class ID string and searches the class list for it. Returns the
    # first matching budget class, or None if nothing was found.
    def get_class(self, class_id):
        for bc in self.classes:
            if bc.match_id(class_id):
                return BudgetResult(success=True, data=bc)
        return BudgetResult(success=False, msg="Couldn't find a match")
    
    # Expects a transaction ID string and searches all classes for it. Returns
    # the transaction object if one is found, or None if nothing is found.
    def get_transaction(self, transaction_id):
        for bc in self.classes:
            for t in bc:
                if t.match_id(transaction_id):
                    return BudgetResult(success=True, data=t)
        return BudgetResult(success=False, msg="Couldn't find a match")

    # Takes in text and searches the expense classes for a matching one.
    # Returns a list of matching BudgetClass objects.
    def search_class(self, text):
        result = []
        # search through all budget classes (sorted)
        for c in self.all():
            # if the class matches the text in one way or another, add it
            if c.match(text):
                result.append(c)
        # build a return object
        succ = len(result) > 0
        m = "" if succ else "Couldn't find any matches"
        return BudgetResult(success=succ, msg=m, data=result)
    
    # Used to search for a transaction given some sort of information about it.
    # Returns a list of transaction objects that matched the text in one way
    # or another. Returns an empty list if no match is found.
    def search_transaction(self, text):
        result = []
        # search through all budget classes (sorted)
        for bc in self.all():
            # search through each transaction in sorted order (most recent
            # will be checked first)
            for t in bc.all():
                if t.match(text):
                    result.append(t)
        # build a return object
        succ = len(result) > 0
        m = "" if succ else "Couldn't find any matches"
        return BudgetResult(success=succ, msg=m, data=result)
    
    # ------------------------------ Removals -------------------------------- #
    # Takes in a budget class and does two things:
    #   1. Deletes its file on disk
    #   2. Removes the class from the Budget object's internal list
    # This throws an exception if the class isn't found within.
    def delete_class(self, bclass):
        # use the given class's ID to find the equivalent object stored in the
        # Budget object, then use it to get the index
        result = self.get_class(bclass.bcid)
        if not result.success:
            return result
        bc = result.data
        idx = self.classes.index(bc)
        self.classes.pop(idx)

        # now, build the file path and delete the file
        sroot = self.save_root_path()
        fpath = os.path.join(sroot, bc.to_file_name())
        os.remove(fpath)
        # attempt to back up
        try:
            bclass.save(self.backup_class_path(bclass))
        except Exception as e:
            m = "Failed to backup class: %s" % e
            return BudgetResult(success=True, msg=m)
        return BudgetResult(success=True)


    # Takes in a transaction and deletes it from its corresponding budget class.
    # Throws an exception if the transaction isn't inside the budget.
    def delete_transaction(self, transaction):
        # locate the true transaction object stored within the budget object AND
        # the true budget class that's storing the transaction
        result = self.get_transaction(transaction.tid)
        if not result.success:
            return result
        t = result.data
        result = self.get_class(transaction.owner.bcid)
        if not result.success:
            return result
        bc = result.data
        
        # remove the transaction from the class, then save the budget class
        bc.remove(t)
        sroot = self.save_root_path()
        fpath = os.path.join(sroot, bc.to_file_name())
        bc.save(fpath)
        # attempt to back up
        try:
            bclass.save(self.backup_class_path(bclass))
        except Exception as e:
            m = "Failed to backup class: %s" % e
            return BudgetResult(success=True, msg=m)
        return BudgetResult(success=True)


    # ---------------------- Manual Saving and Backups ----------------------- #
    # Takes in a class and saves it to the correct location.
    def update_class(self, bclass):
        # first, delete the old version of the same budget class. Then, save the
        # new one with its updated fields
        result = self.delete_class(bclass)
        if not result.success:
            return result
        result = self.add_class(bclass)
        if not result.success:
            return result
        return BudgetResult(success=True)

    # Takes in a timestamp (datetime.now() by default) and uses it to determine
    # the current reset date, and from it, a path to the directory into which a
    # budget object should be saved. If the directory doesn't exist, this
    # function also creates it.
    def save_root_path(self, dt=datetime.now()):
        # using the given datetime, we'll iterate through the reset dates and
        # determine which one to use for the directory path
        d = self.conf.reset_dates[0]
        best_score = None
        for rd in self.conf.reset_dates:
            # compute a difference between the month and the day. We'll select
            # the reset date with the smallest positive value as the most
            # recently-passed
            diff = (dt.month - rd.month) + ((dt.day - rd.day) / 40.0)
            if diff >= 0.0 and (best_score == None or diff < best_score):
                best_score = diff
                d = rd
        
        # construct the folder directory path, create it if necessary, and
        # return it
        dpath = self.conf.save_location + "/%d-%d-%d" % (dt.year, d.month, d.day)
        assert not os.path.isfile(dpath), "save root path is a file: %s" % dpath
        if not os.path.isdir(dpath):
            os.mkdir(dpath)
        return dpath

    # Attempts to set up the current backup location based on the config's
    # 'backup_location' entry. Returns the path to the backup directory.
    def backup_setup(self, dt=datetime.now()):
        # we'll compute the same kind of score we do above in 'save_root_path'
        # to determine which reset date to use for backup
        d = self.conf.reset_dates[0]
        best_score = None
        for rd in self.conf.reset_dates:
            diff = (dt.month - rd.month) + ((dt.day - rd.day) / 40.0)
            if diff >= 0.0 and (best_score == None or diff < best_score):
                best_score = diff
                d = rd
        
        # with the correct date, build a backup file path
        dpath = "%s/%d-%d-%d" % (self.conf.backup_location, dt.year,
                                 d.month, d.day)

        # if the directory doesn't exist, create it
        assert not os.path.isfile(dpath), "backup location is a file"
        if not os.path.isdir(dpath):
            os.mkdir(dpath)
        return dpath
    
    # Takes in a budget class and returns the path at which it should be saved.
    def backup_class_path(self, bclass, dt=datetime.now()):
        dpath = self.backup_setup(dt=dt)
        fpath = os.path.join(dpath, bclass.to_file_name())
        return fpath
    
    # Takes in a file path and attempts to create an Excel file for the entire
    # budget.
    def write_to_excel(self, fpath):
        # first, create an Excel workbook object, and create a worksheet for
        # each budget class
        wb = Workbook()

        # ------------------------ Overview Worksheet ------------------------ #
        # get the active sheet, change its name, and set a few fields
        ws1 = wb.active
        ws1.title = "Overview"
        ws1["A1"] = "Name"
        ws1["B1"] = self.conf.name
        ws1["A2"] = "Start Date"
        start_date = self.reset_dates[-1]
        start_date = start_date.replace(year=start_date.year - 1)
        ws1["B2"] = start_date
        ws1["B2"].number_format = "yyyy-mm-dd"
        ws1["A3"] = "End Date"
        end_date = self.reset_dates[0]
        end_date = datetime.fromtimestamp(end_date.timestamp() - 86400)
        ws1["B3"] = end_date
        ws1["B3"].number_format = "yyyy-mm-dd"
        
        # quickly loop through the classes together to compute totals for income
        # and expenses
        itotal = 0.0
        etotal = 0.0
        for bc in self.classes:
            # iterate through the class' history
            total = 0.0
            for t in bc.history:
                total += t.price
            # depending on the type, add to the correct counter
            if bc.ctype == BudgetClassType.INCOME:
                itotal += total
            else:
                etotal += total

        # set more overview fields
        ws1["A4"] = "Total Income"
        ws1["B4"] = itotal
        ws1["B4"].number_format = "$#0.00"
        ws1["A5"] = "Total Expenses"
        ws1["B5"] = etotal
        ws1["B5"].number_format = "$#0.00"
        ws1["A6"] = "Surplus"
        ws1["B6"] = itotal - etotal
        ws1["B6"].number_format = "$#0.00"

        # set up a header for a table we'll build below
        ws1["A8"] = "Budget Class"
        ws1["B8"] = "Type"
        ws1["C8"] = "Total"
        ws1["D8"] = "Target"
        row_num = 9
        
        # set column widths
        ws1.column_dimensions["A"].width = 35
        ws1.column_dimensions["B"].width = 20
        # set a few fonts
        header_font = Font(bold=True)
        for cell in ["A1", "A2", "A3", "A4", "A5", "A6", "A8", "B8", "C8", "D8"]:
            ws1[cell].font = header_font

        # ----------------------- Per-Class Worksheet ------------------------ #
        # sort all classes first by type, then by name
        classes = sorted(self.classes, key=lambda bc: (int(bc.ctype), bc.name.lower()))
        bccount = 0
        for bc in classes:
            ws = wb.create_sheet(title=bc.name)
            # set the color appropriately
            if bc.ctype == BudgetClassType.INCOME:
                ws.sheet_properties.tabColor = "A9D08E"
            else:
                ws.sheet_properties.tabColor = "FFD966"

            # give our table some titles
            ws["A1"] = "Date"
            ws["A1"].font = header_font
            ws["B1"] = "Price"
            ws["B1"].font = header_font
            ws["C1"] = "Vendor"
            ws["C1"].font = header_font
            ws["D1"] = "Description"
            ws["D1"].font = header_font

            # set cell sizes
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 40

            # sort the transactions in ascending order by date
            ts = sorted(bc.history, key=lambda t: t.timestamp.timestamp())

            # for each transaction in the budget, we'll get its JSON form and
            # use it to construct a row
            idx = 2
            total = 0.0
            for t in ts:
                jdata = t.to_json()
                total += jdata["price"]
                # set row values
                ws["A%d" % idx] = datetime.fromtimestamp(jdata["timestamp"])
                ws["A%d" % idx].number_format = "yyyy-mm-dd"
                ws["B%d" % idx] = jdata["price"]
                ws["B%d" % idx].number_format = "$#0.00"
                ws["C%d" % idx] = jdata["vendor"]
                ws["D%d" % idx] = jdata["description"]
                # increment counters
                idx += 1

            # add a row to our overview worksheet
            tstr = "INCOME" if bc.ctype == BudgetClassType.INCOME else "EXPENSE"
            rn = str(row_num)
            ws1["A" + rn] = bc.name
            ws1["B" + rn] = tstr
            ws1["C" + rn] = total
            ws1["C" + rn].number_format = "$#0.00"
            if bc.target != None:
                ws1["D" + rn] = bc.target.get_value(total_income=itotal)
                ws1["D" + rn].number_format = "$#0.00"
            row_num += 1
                
        # ----------------------------- Savings ------------------------------ #
        # now we'll add savings information to the overview sheet
        row_num += 1
        rn = str(row_num)
        ws1["A" + rn] = "Savings Category"
        ws1["B" + rn] = "Percentage of Surplus"
        ws1["C" + rn] = "Amount to Save"
        for c in ["A", "B", "C"]:
            ws1[c + rn].font = header_font
        # iterate through each savings category
        for sc in sorted(self.savings, key=lambda sc: sc.name):
            row_num += 1
            rn = str(row_num)
            ws1["A" + rn] = sc.name
            ws1["B" + rn] = sc.percent
            ws1["B" + rn].number_format = "%#0"
            ws1["C" + rn] = sc.percent * max(0.0, itotal - etotal)
            ws1["C" + rn].number_format = "$#0.00"
        
        # save the workbook
        wb.save(filename=fpath)

    # ---------------------------- Other Helpers ----------------------------- #
    # Returns *all* budget classes within the budget, in sorted order by name.
    def all(self):
        return sorted(self.classes, key=lambda bc: bc.name.lower())
    
    # Creates one monolithic JSON struct containing all budget classes and
    # all transactions within them.
    def to_json(self):
        jdata = []
        # iterate through all classes and add their JSON to the data
        for c in self.classes:
            jdata.append(c.to_json())
        return jdata
    
    # Returns the number of seconds from now until the next scheduled budget
    # reset.
    def time_to_reset(self):
        # get the earliest reset date and compare it against the current time
        nrd = self.conf.reset_dates[0]
        now = datetime.now()
        return nrd.timestamp() - now.timestamp()


